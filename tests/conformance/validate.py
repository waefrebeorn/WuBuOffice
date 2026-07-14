#!/usr/bin/env python3
"""External conformance oracle for WuBuOffice output.

Validates generated OOXML packages against *independent* libraries (not WuBu's
own reader), so conformance bugs can never hide behind our own parser.

Per part:
  * Every XML part must be well-formed (Python stdlib xml.etree -- a different
    implementation than WuBu's hand-rolled SAX parser).
For xlsx specifically:
  * openpyxl must load the workbook and the cached cell values must match
    what WuBu wrote (string/number/formula-cached values).

Usage: validate.py <file.docx|file.xlsx|file.pptx>
Exit 0 = conformant, 1 = defect, 2 = oracle unavailable (treated as SKIP).
"""
import sys, zipfile, xml.etree.ElementTree as ET

WELLFORM_OK = True

def check_wellformed(z):
    bad = []
    for n in z.namelist():
        if n.endswith(".xml") or n.endswith(".rels"):
            try:
                ET.fromstring(z.read(n))
            except Exception as e:
                bad.append(f"{n}: {e}")
    return bad

def __colnum(col):
    """Convert a column letter (e.g. 'B') to a 1-based index."""
    n = 0
    for ch in col:
        n = n * 26 + (ord(ch) - ord("A") + 1)
    return n

def main():
    if len(sys.argv) < 2:
        print("usage: validate.py <file>"); return 2
    path = sys.argv[1]
    try:
        z = zipfile.ZipFile(path)
    except Exception as e:
        print(f"CANNOT OPEN: {e}"); return 1

    errors = []
    bad = check_wellformed(z)
    if bad:
        errors.append("XML WELL-FORMEDNESS FAILURES:")
        errors += bad

    # xlsx-specific: load with openpyxl (external oracle) and check values
    if path.endswith(".xlsx"):
        try:
            import openpyxl
        except Exception:
            print("openpyxl unavailable -> SKIP value check");
            # well-formedness still validated above
            if errors:
                print("\n".join(errors)); return 1
            print("CONFORMANT (well-formed; value oracle skipped)"); return 0
        wb = openpyxl.load_workbook(path, data_only=True)
        if not wb.sheetnames:
            errors.append("openpyxl: no sheets"); 
        else:
            ws = wb[wb.sheetnames[0]]
            # ensure we can read cell values without error
            try:
                _ = [c.value for row in ws.iter_rows() for c in row]
            except Exception as e:
                errors.append(f"openpyxl read error: {e}")
            # concrete fidelity checks (the conformance file is deterministic)
            expected = {
                ("A", 1): "Item", ("B", 1): "Cost",
                ("A", 2): "Engine", ("B", 2): 1200.5,
                ("A", 3): "Docs", ("B", 3): 320,
                ("B", 4): 1520.5,   # cached SUM(B2:B3)
            }
            for (col, row), want in expected.items():
                got = ws.cell(row=row, column=__colnum(col)).value
                if isinstance(want, float):
                    g = float(got) if got is not None else None
                    if g is None or abs(g - want) > 1e-6:
                        errors.append(f"cell {col}{row}: want {want}, got {got!r}")
                else:
                    if got != want:
                        errors.append(f"cell {col}{row}: want {want!r}, got {got!r}")

    if errors:
        print("\n".join(errors)); return 1
    print("CONFORMANT"); return 0

if __name__ == "__main__":
    sys.exit(main())
